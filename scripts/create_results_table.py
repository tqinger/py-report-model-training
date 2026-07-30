from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from tcm_qwen_eval.dataset import ACTUAL_TRAINING_DATA_DIR, read_selection
from tcm_qwen_eval.dynamic_input import extract_dynamic_input, task_label

MODELS = ("Qwen/Qwen3-0.6B", "Qwen/Qwen3-1.7B", "Qwen/Qwen3-4B")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _read_generations(path: Path) -> dict[tuple[str, str], str]:
    """Read model outputs from the original longitudinal baseline workbook."""
    workbook = load_workbook(path, data_only=False, read_only=True)
    samples = workbook.worksheets[1]
    generations: dict[tuple[str, str], str] = {}
    for row in samples.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        model, example_id, output = row[0], row[3], row[8]
        generations[(model, example_id)] = output
    workbook.close()
    return generations


def _fit_columns(sheet) -> None:
    widths = {"A": 24, "B": 72, "C": 60, "D": 55, "E": 14, "F": 55, "G": 14, "H": 55, "I": 14, "J": 40}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_results_table(
    output: Path,
    selected_examples,
    generations: dict[tuple[str, str], str],
    source_data_dir: Path,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.title = "结果表"
    sheet.append(
        [
            "任务类型",
            "数据源",
            "动态输入数据",
            "Qwen3-0.6B 输出",
            "Qwen3-0.6B 评分",
            "Qwen3-1.7B 输出",
            "Qwen3-1.7B 评分",
            "Qwen3-4B 输出",
            "Qwen3-4B 评分",
            "备注",
        ]
    )
    for example in selected_examples:
        outputs = []
        for model in MODELS:
            value = generations.get((model, example.id))
            if not value:
                raise ValueError(f"Missing generated output for {model} / {example.id}")
            outputs.extend((value, None))
        source = source_data_dir / f"{example.id}.json"
        sheet.append([task_label(example), str(source), extract_dynamic_input(example), *outputs, None])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    _fit_columns(sheet)
    book.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a single side-by-side LLM results table.")
    parser.add_argument(
        "--selection",
        type=Path,
        default=Path("artifacts/baseline_selection_report_system.json"),
    )
    parser.add_argument("--baseline", type=Path, default=Path("artifacts/qwen3_baseline.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("artifacts/qwen3_results_table.xlsx"))
    parser.add_argument("--source-data-dir", type=Path, default=ACTUAL_TRAINING_DATA_DIR)
    args = parser.parse_args()

    write_results_table(
        args.output,
        read_selection(args.selection),
        _read_generations(args.baseline),
        args.source_data_dir,
    )
    print(f"Results table written to {args.output}")


if __name__ == "__main__":
    main()
