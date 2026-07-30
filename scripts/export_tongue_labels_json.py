"""Export the ``舌`` worksheet from the label workbook as structured JSON.

The source worksheet uses merged cells for types and sometimes continues a
label's differential diagnoses on rows whose label cell is blank.  This
exporter resolves those presentation details into a stable, machine-readable
representation.
"""

from __future__ import annotations

import argparse
import json
from itertools import combinations as iter_combinations
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_SOURCE_DIR = Path(r"D:\shanjiyun\data")
DEFAULT_SOURCE_SUFFIX = "2026.0714.xlsx"
DEFAULT_OUTPUT = Path("outputs/tongue_labels.json")

TYPE_SLUGS = {
    "舌神": "spirit",
    "舌色": "color",
    "舌形": "shape",
    "舌苔": "coating",
    "苔质-薄厚": "coating_thickness",
    "苔质-润燥": "coating_moisture",
    "苔质-腐腻": "coating_texture",
    "苔色": "coating_color",
    "舌下络脉": "sublingual_veins",
    "舌态": "movement",
}

TONGUE_COLOR = "\u820c\u8272"
TONGUE_SHAPE = "\u820c\u5f62"
TONGUE_COATING = "\u820c\u82d4"

EXCLUDED_COLOR_LABELS = {
    "\u6de1\u7d2b\u820c",
    "\u7edb\u7d2b\u820c",
    "\u7d2b\u6697\u820c",
    "\u7d2b\u6591\u820c",
}
EXCLUDED_SHAPE_LABELS = {"\u8001\u820c", "\u5ae9\u820c"}
NORMAL_SHAPE_LABEL = "\u6b63\u5e38\u820c\u5f62"
EXCLUDED_COATING_LABELS = {
    "\u82b1\u5265\u82d4",
    "\u5730\u56fe\u820c",
    "\u524d\u5265\u8131\u82d4",
    "\u4e2d\u5265\u8131\u82d4",
    "\u6839\u5265\u8131\u82d4",
}
SPECIAL_COATING_LABELS = {
    "\u82b1\u5265\u82d4",
    "\u5730\u56fe\u820c",
    "\u82b1\u5265\u82d4(\u5730\u56fe\u820c)",
}


def find_source_workbook(source_dir: Path, source_suffix: str) -> Path:
    """Find one non-lock workbook matching the configured suffix."""

    candidates = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$") and path.name.endswith(source_suffix)
    )
    if len(candidates) != 1:
        raise ValueError(
            f"Expected exactly one source workbook ending with {source_suffix!r} in {source_dir}, "
            f"found {len(candidates)}."
        )
    return candidates[0]


def export_tongue_sheet(source_path: Path) -> dict[str, Any]:
    """Convert the source workbook's ``舌`` worksheet to the agreed schema."""

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if "舌" not in workbook.sheetnames:
        raise ValueError(f"Worksheet '舌' is missing from {source_path}.")

    worksheet = workbook["舌"]
    labels: list[dict[str, Any]] = []
    type_label_ids: dict[str, list[str]] = {}
    type_counts: dict[str, int] = {}
    current_type: str | None = None
    current_label: dict[str, Any] | None = None

    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row = list(values)
        type_cell = row[0] if len(row) > 0 else None
        label_cell = row[1] if len(row) > 1 else None
        conclusion_cell = row[2] if len(row) > 2 else None
        description_cell = row[3] if len(row) > 3 else None

        if type_cell is not None and str(type_cell).strip():
            current_type = str(type_cell).strip()
            type_label_ids.setdefault(current_type, [])
            type_counts.setdefault(current_type, 0)

        label = str(label_cell).strip() if label_cell is not None else ""
        conclusion = str(conclusion_cell).strip() if conclusion_cell is not None else ""
        description = str(description_cell).strip() if description_cell is not None else ""
        if not any((label, conclusion, description)):
            continue

        if label:
            if current_type is None:
                raise ValueError(f"Row {row_number} has a label but no inherited type.")

            type_counts[current_type] += 1
            label_id = f"tongue_{TYPE_SLUGS.get(current_type, 'unknown')}_{type_counts[current_type]:03d}"
            current_label = {
                "id": label_id,
                "type": current_type,
                "label": label,
                "diagnoses": [],
            }
            labels.append(current_label)
            type_label_ids[current_type].append(label_id)
        elif current_label is None:
            raise ValueError(f"Row {row_number} has diagnosis data but no preceding label.")

        if not conclusion or not description:
            raise ValueError(
                f"Row {row_number} must contain both a conclusion and description for its label diagnosis."
            )

        current_label["diagnoses"].append(
            {
                "conclusion": conclusion,
                "description": description,
            }
        )

    if not labels:
        raise ValueError("No tongue labels were exported.")

    payload = {
        "source": {
            "workbook": source_path.name,
            "worksheet": "舌",
            "remark_column_included": False,
        },
        "tongue_types": [
            {"name": type_name, "label_ids": label_ids}
            for type_name, label_ids in type_label_ids.items()
        ],
        "labels": labels,
    }
    payload["combinations"] = build_combinations(labels)
    return payload


def build_combinations(labels: list[dict[str, Any]]) -> list[list[str]]:
    """Generate every allowed complete color, shape, and coating combination."""

    labels_by_type: dict[str, list[str]] = {}
    for item in labels:
        labels_by_type.setdefault(item["type"], []).append(item["label"])

    color_options = [
        label for label in labels_by_type[TONGUE_COLOR] if label not in EXCLUDED_COLOR_LABELS
    ]
    selectable_shapes = [
        label
        for label in labels_by_type[TONGUE_SHAPE]
        if label not in EXCLUDED_SHAPE_LABELS and label != NORMAL_SHAPE_LABEL
    ]
    shape_options = [[NORMAL_SHAPE_LABEL]]
    shape_options.extend(
        list(selection)
        for size in range(1, len(selectable_shapes) + 1)
        for selection in iter_combinations(selectable_shapes, size)
    )

    selectable_coatings = [
        label
        for label in labels_by_type[TONGUE_COATING]
        if label not in EXCLUDED_COATING_LABELS
    ]
    special_coatings = [
        label for label in selectable_coatings if label in SPECIAL_COATING_LABELS
    ]
    ordinary_coatings = [
        label for label in selectable_coatings if label not in SPECIAL_COATING_LABELS
    ]
    coating_options = [[label] for label in ordinary_coatings]
    for special in special_coatings:
        coating_options.append([special])
        coating_options.extend([special, ordinary] for ordinary in ordinary_coatings)

    return [
        [color, *shape, *coating]
        for color in color_options
        for shape in shape_options
        for coating in coating_options
    ]


def validate_export(payload: dict[str, Any]) -> None:
    """Validate the two JSON views before writing them to disk."""

    labels = payload["labels"]
    label_ids = [item["id"] for item in labels]
    if len(label_ids) != len(set(label_ids)):
        raise ValueError("Generated label IDs are not unique.")

    label_id_set = set(label_ids)
    indexed_ids = [
        label_id
        for type_item in payload["tongue_types"]
        for label_id in type_item["label_ids"]
    ]
    if len(indexed_ids) != len(set(indexed_ids)) or set(indexed_ids) != label_id_set:
        raise ValueError("Type label index does not match the flattened labels.")

    for label in labels:
        if not label["type"] or not label["label"] or not label["diagnoses"]:
            raise ValueError(f"Label {label['id']} is incomplete.")
        for diagnosis in label["diagnoses"]:
            if not diagnosis["conclusion"] or not diagnosis["description"]:
                raise ValueError(f"Label {label['id']} contains an incomplete diagnosis.")
            if "explanations" in diagnosis:
                raise ValueError(f"Label {label['id']} unexpectedly includes explanations.")

    labels_by_type: dict[str, set[str]] = {}
    for label in labels:
        labels_by_type.setdefault(label["type"], set()).add(label["label"])

    allowed_colors = labels_by_type[TONGUE_COLOR] - EXCLUDED_COLOR_LABELS
    allowed_shapes = labels_by_type[TONGUE_SHAPE] - EXCLUDED_SHAPE_LABELS
    allowed_coatings = labels_by_type[TONGUE_COATING] - EXCLUDED_COATING_LABELS
    combinations = payload["combinations"]
    if len(combinations) != 10_176 or len({tuple(item) for item in combinations}) != 10_176:
        raise ValueError("Combination count or uniqueness does not match the configured rules.")

    for combination in combinations:
        if not isinstance(combination, list) or not combination:
            raise ValueError("Every combination must be a non-empty list.")
        colors = [item for item in combination if item in allowed_colors]
        shapes = [item for item in combination if item in allowed_shapes]
        coatings = [item for item in combination if item in allowed_coatings]
        if len(colors) != 1 or not shapes or not coatings:
            raise ValueError(f"Combination is missing a required type: {combination!r}")
        if len(colors) + len(shapes) + len(coatings) != len(combination):
            raise ValueError(f"Combination includes an invalid label: {combination!r}")
        if NORMAL_SHAPE_LABEL in shapes and len(shapes) != 1:
            raise ValueError(f"Normal shape cannot be combined with other shapes: {combination!r}")
        special_count = sum(item in SPECIAL_COATING_LABELS for item in coatings)
        ordinary_count = len(coatings) - special_count
        if special_count > 1 or ordinary_count > 1:
            raise ValueError(f"Combination violates coating exclusivity: {combination!r}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export the tongue-label worksheet as JSON.")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--source-suffix", default=DEFAULT_SOURCE_SUFFIX)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    source_path = find_source_workbook(args.source_dir, args.source_suffix)
    payload = export_tongue_sheet(source_path)
    validate_export(payload)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    diagnosis_count = sum(len(label["diagnoses"]) for label in payload["labels"])
    print(
        f"Exported {len(payload['labels'])} labels and {diagnosis_count} diagnoses "
        f"across {len(payload['tongue_types'])} types to {args.output}."
    )


if __name__ == "__main__":
    main()
