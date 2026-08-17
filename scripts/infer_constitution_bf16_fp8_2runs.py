"""Generate two BF16 and FP8 responses for five held-out constitution cases.

Each adapter is loaded independently, merged into Qwen3-4B, and evaluated twice
per prompt with fixed, distinct random seeds.  FP8 uses TorchAO dynamic-activation
and FP8-weight linear-layer quantization after the LoRA adapter is merged.
"""

from __future__ import annotations

import argparse
import gc
import json
import random
import time
from dataclasses import asdict
from pathlib import Path
from typing import Any

import peft.tuners.lora.torchao as peft_torchao
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from peft import PeftModel
from torchao.quantization import float8_dynamic_activation_float8_weight, quantize_
from transformers import AutoModelForCausalLM, AutoTokenizer

from tcm_qwen_eval.dataset import Example, load_examples
from tcm_qwen_eval.tongue_qlora import split_constitution_examples

DEFAULT_BASE_MODEL = Path(
    "artifacts/hf_cache/hub/models--Qwen--Qwen3-4B/snapshots/"
    "1cfa9a7208912126459214e8b04321603b3df60c"
)
DEFAULT_ADAPTERS = (
    Path("artifacts/qwen3-4b-constitution-qlora/checkpoint-3486"),
    Path("artifacts/qwen3-4b-constitution-qlora/checkpoint-2324"),
)
DEFAULT_OUTPUT_DIR = Path("outputs/constitution_4b_checkpoints_bf16_fp8_5cases_2runs_20260812")
SPLIT_SEED = 20260811
SELECTION_SEED = 20260812
GENERATION_SEEDS = (2026081201, 2026081202)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run BF16/FP8 constitution inference twice for five held-out cases."
    )
    parser.add_argument("--data-dir", type=Path, default=Path("data/constitution-analysis"))
    parser.add_argument("--base-model", type=Path, default=DEFAULT_BASE_MODEL)
    parser.add_argument("--adapter", type=Path, action="append", default=None)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--max-new-tokens", type=int, default=384)
    parser.add_argument("--temperature", type=float, default=0.7)
    parser.add_argument("--top-p", type=float, default=0.8)
    parser.add_argument(
        "--checkpoint-index",
        type=int,
        action="append",
        choices=(1, 2),
        default=None,
        help="Run only checkpoint 1 and/or 2; supports staged execution.",
    )
    return parser.parse_args()


def select_cases(data_dir: Path) -> list[Example]:
    """Select five deterministic, source-distinct examples from the held-out split."""
    test_examples = split_constitution_examples(load_examples(data_dir), SPLIT_SEED)["test"]
    by_group: dict[str, Example] = {}
    for example in sorted(test_examples, key=lambda item: item.id):
        by_group.setdefault(example.group_id, example)
    if len(by_group) < 5:
        raise ValueError(f"Expected at least five held-out source groups, found {len(by_group)}")
    chooser = random.Random(SELECTION_SEED)
    return [by_group[group_id] for group_id in chooser.sample(sorted(by_group), 5)]


def constitution_info(example: Example) -> str:
    marker = "体质辨识信息："
    if marker not in example.user:
        raise ValueError(f"{example.id}: missing {marker!r}")
    return example.user.split(marker, maxsplit=1)[1].strip()


def load_merged_model(base_model_path: Path, adapter_path: Path) -> tuple[Any, AutoTokenizer]:
    tokenizer = AutoTokenizer.from_pretrained(adapter_path, local_files_only=True)
    if tokenizer.pad_token_id is None:
        tokenizer.pad_token = tokenizer.eos_token
    base = AutoModelForCausalLM.from_pretrained(
        base_model_path,
        torch_dtype=torch.bfloat16,
        local_files_only=True,
        low_cpu_mem_usage=True,
    ).to("cuda")
    print("Base model moved to CUDA; loading LoRA adapter...", flush=True)
    # This project pins PEFT 0.20 with TorchAO 0.11.  Inject LoRA through the
    # standard linear path, then quantize only after it has been merged.
    peft_torchao.is_torchao_available = lambda: False
    merged = PeftModel.from_pretrained(base, adapter_path, local_files_only=True).merge_and_unload()
    merged.eval()
    return merged, tokenizer


@torch.inference_mode()
def generate(
    model: Any,
    tokenizer: AutoTokenizer,
    messages: list[dict[str, str]],
    seed: int,
    max_new_tokens: int,
    temperature: float,
    top_p: float,
) -> str:
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    inputs = tokenizer.apply_chat_template(
        messages[:2],
        tokenize=True,
        add_generation_prompt=True,
        enable_thinking=False,
        return_dict=True,
        return_tensors="pt",
    ).to("cuda")
    generated = model.generate(
        **inputs,
        do_sample=True,
        temperature=temperature,
        top_p=top_p,
        max_new_tokens=max_new_tokens,
        pad_token_id=tokenizer.pad_token_id,
        eos_token_id=tokenizer.eos_token_id,
    )
    output_ids = generated[0, inputs.input_ids.shape[1] :]
    return tokenizer.decode(output_ids, skip_special_tokens=True).strip()


def predict_variant(
    model: Any,
    tokenizer: AutoTokenizer,
    cases: list[Example],
    precision: str,
    max_new_tokens: int,
    temperature: float,
    top_p: float,
) -> list[list[dict[str, Any]]]:
    rows: list[list[dict[str, Any]]] = []
    for case_index, example in enumerate(cases, start=1):
        answers: list[dict[str, Any]] = []
        for answer_index, seed in enumerate(GENERATION_SEEDS, start=1):
            started = time.perf_counter()
            output = generate(
                model,
                tokenizer,
                example.messages,
                seed,
                max_new_tokens,
                temperature,
                top_p,
            )
            answers.append(
                {
                    "answer_index": answer_index,
                    "seed": seed,
                    "output": output,
                    "output_chars": len(output),
                    "latency_seconds": round(time.perf_counter() - started, 3),
                }
            )
            print(
                f"{precision}: case {case_index}/{len(cases)}, answer {answer_index}/2 complete",
                flush=True,
            )
        rows.append(answers)
    return rows


def add_summary_sheet(workbook: Workbook, payload: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "说明"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110
    values = [
        ("报告说明", "Qwen3-4B 体质分析 LoRA：2 个 checkpoint、BF16 与 FP8 各 5 个案例，每例生成 2 份答案。"),
        ("案例来源", "按来源分组的独立测试集；随机种子 20260812，5 个案例彼此来源不同。"),
        ("生成方式", "采样生成（temperature=0.7，top_p=0.8），每例答案 1/2 使用固定不同随机种子，便于复现。"),
        ("FP8 方式", "合并 LoRA 后使用 TorchAO：动态激活 + FP8 权重线性层量化。"),
        ("基座模型", payload["base_model"]),
        ("输出总数", "40 条：2 checkpoint × 2 精度 × 5 案例 × 2 答案。"),
    ]
    for row in values:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[index].height = 40


def add_checkpoint_sheet(
    workbook: Workbook,
    checkpoint_name: str,
    cases: list[Example],
    checkpoint_results: dict[str, list[list[dict[str, Any]]]],
) -> None:
    sheet = workbook.create_sheet(checkpoint_name)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C2"
    headers = [
        "案例",
        "体质辨识信息",
        "BF16 答案 1",
        "BF16 答案 2",
        "FP8 答案 1",
        "FP8 答案 2",
    ]
    sheet.append(headers)
    for case_index, example in enumerate(cases, start=1):
        bf16_answers = checkpoint_results["bf16"][case_index - 1]
        fp8_answers = checkpoint_results["fp8"][case_index - 1]
        sheet.append(
            [
                f"案例 {case_index}\n{example.id}",
                constitution_info(example),
                bf16_answers[0]["output"],
                bf16_answers[1]["output"],
                fp8_answers[0]["output"],
                fp8_answers[1]["output"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=6):
        for cell in row:
            cell.alignment = body_alignment
            cell.border = border
    widths = (34, 44, 72, 72, 72, 72)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 32
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 230
    sheet.auto_filter.ref = f"A1:F{sheet.max_row}"


def write_workbook(output_path: Path, payload: dict[str, Any]) -> None:
    workbook = Workbook()
    add_summary_sheet(workbook, payload)
    cases = payload["cases"]
    case_examples = [
        Example(
            id=case["id"],
            domain="constitution-analysis",
            task="constitution_combined_analysis",
            group_id=case["group_id"],
            messages=case["messages"],
        )
        for case in cases
    ]
    for checkpoint in payload["checkpoints"]:
        add_checkpoint_sheet(
            workbook,
            checkpoint["name"],
            case_examples,
            checkpoint["results"],
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    if not torch.cuda.is_available():
        raise RuntimeError("CUDA is unavailable; BF16/FP8 inference requires a CUDA GPU.")
    if not args.base_model.is_dir():
        raise FileNotFoundError(f"Base model not found: {args.base_model}")

    adapters = tuple(args.adapter) if args.adapter else DEFAULT_ADAPTERS
    if len(adapters) != 2:
        raise ValueError("Provide exactly two --adapter values.")
    for adapter in adapters:
        if not adapter.is_dir():
            raise FileNotFoundError(f"Adapter not found: {adapter}")
    if args.checkpoint_index:
        adapters = tuple(adapters[index - 1] for index in args.checkpoint_index)

    cases = select_cases(args.data_dir)
    json_path = args.output_dir / "prediction_results.json"
    if json_path.exists():
        payload: dict[str, Any] = json.loads(json_path.read_text(encoding="utf-8"))
        existing_names = {item["name"] for item in payload["checkpoints"]}
        if payload["cases"] != [asdict(case) for case in cases]:
            raise ValueError("Existing output uses a different case selection; choose a new output directory.")
    else:
        payload = {
        "base_model": str(args.base_model),
        "data_dir": str(args.data_dir),
        "case_selection": {
            "split": "held-out test split, source-grouped 80/10/10",
            "split_seed": SPLIT_SEED,
            "selection_seed": SELECTION_SEED,
        },
        "generation": {
            "do_sample": True,
            "temperature": args.temperature,
            "top_p": args.top_p,
            "max_new_tokens": args.max_new_tokens,
            "thinking": False,
            "seeds": list(GENERATION_SEEDS),
        },
        "fp8_quantization": "TorchAO dynamic-activation FP8-weight linear layers after LoRA merge",
        "cases": [asdict(case) for case in cases],
            "checkpoints": [],
        }
        existing_names = set()

    for adapter in adapters:
        if adapter.name in existing_names:
            print(f"Skipping already completed {adapter.name}.", flush=True)
            continue
        print(f"Loading and merging {adapter}...", flush=True)
        model, tokenizer = load_merged_model(args.base_model, adapter)
        print(f"Running BF16 for {adapter.name}...", flush=True)
        bf16 = predict_variant(
            model, tokenizer, cases, "BF16", args.max_new_tokens, args.temperature, args.top_p
        )
        print(f"Quantizing {adapter.name} to FP8...", flush=True)
        quantize_(model, float8_dynamic_activation_float8_weight())
        model.eval()
        print(f"Running FP8 for {adapter.name}...", flush=True)
        fp8 = predict_variant(
            model, tokenizer, cases, "FP8", args.max_new_tokens, args.temperature, args.top_p
        )
        payload["checkpoints"].append(
            {"name": adapter.name, "adapter": str(adapter), "results": {"bf16": bf16, "fp8": fp8}}
        )
        args.output_dir.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        write_workbook(args.output_dir / "constitution_bf16_fp8_5cases_2answers.xlsx", payload)
        print(f"Saved checkpoint results for {adapter.name}.", flush=True)
        del model
        gc.collect()
        torch.cuda.empty_cache()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    workbook_path = args.output_dir / "constitution_bf16_fp8_5cases_2answers.xlsx"
    write_workbook(workbook_path, payload)
    print(json.dumps({"json": str(json_path), "workbook": str(workbook_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
