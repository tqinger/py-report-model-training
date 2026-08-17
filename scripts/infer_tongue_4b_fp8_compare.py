"""Run FP8 inference for the saved 4B tongue-model sample and write a BF16/FP8 comparison.

The source JSON supplies both the fixed prompts and the existing BF16 outputs.  The
4B LoRA adapter is merged into its base model before TorchAO applies dynamic-
activation, FP8-weight quantization to all linear layers.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import peft.tuners.lora.torchao as peft_torchao
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from peft import PeftModel
from torchao.quantization import float8_dynamic_activation_float8_weight, quantize_
from transformers import AutoModelForCausalLM, AutoTokenizer

DEFAULT_SOURCE = Path(
    "outputs/tongue_models_fresh_10sample_20260804/inference_results.json"
)
DEFAULT_OUTPUT_DIR = Path("outputs/tongue_4b_bf16_fp8_comparison_20260810")
DEFAULT_BASE_MODEL = Path(
    "artifacts/hf_cache/hub/models--Qwen--Qwen3-4B/snapshots/"
    "1cfa9a7208912126459214e8b04321603b3df60c"
)
DEFAULT_ADAPTER = Path("artifacts/qwen3-4b-tongue-conversations-qlora")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare saved 4B BF16 outputs with FP8 inference on identical prompts."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--base-model", type=Path, default=DEFAULT_BASE_MODEL)
    parser.add_argument("--adapter", type=Path, default=DEFAULT_ADAPTER)
    parser.add_argument("--max-new-tokens", type=int, default=256)
    parser.add_argument(
        "--rerun-bf16",
        action="store_true",
        help="Generate the BF16 baseline in this run instead of reusing the saved output.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Use only the first N saved prompts (useful for a one-prompt precision check).",
    )
    parser.add_argument(
        "--include-system",
        action="store_true",
        help="Use the source conversation's training-time system + user messages.",
    )
    return parser.parse_args()


def get_4b_records(source: Path) -> list[dict[str, object]]:
    payload = json.loads(source.read_text(encoding="utf-8"))
    for model in payload["models"]:
        if model["model"] == "Qwen/Qwen3-4B":
            return model["records"]
    raise ValueError(f"No Qwen/Qwen3-4B records found in {source}")


def tongue_features(dynamic_input: str) -> str:
    marker = "舌象特征："
    if marker not in dynamic_input:
        raise ValueError(f"Input does not contain {marker!r}: {dynamic_input!r}")
    return dynamic_input.split(marker, maxsplit=1)[1].strip()


def prompt_messages(record: dict[str, object], include_system: bool) -> list[dict[str, str]]:
    user = str(record["dynamic_input"])
    if not include_system:
        return [{"role": "user", "content": user}]

    example_stem = str(record["example_id"]).rsplit("/", maxsplit=1)[-1]
    source_path = Path("data/conversations") / f"{example_stem}.json"
    source_messages = json.loads(source_path.read_text(encoding="utf-8"))["messages"][:2]
    if [message["role"] for message in source_messages] != ["system", "user"]:
        raise ValueError(f"{source_path}: expected system + user messages")
    if source_messages[1]["content"] != user:
        raise ValueError(f"{source_path}: user message does not match the fixed inference prompt")
    return source_messages


@torch.inference_mode()
def generate(
    model: torch.nn.Module,
    tokenizer: AutoTokenizer,
    messages: list[dict[str, str]],
    max_new_tokens: int,
) -> str:
    inputs = tokenizer.apply_chat_template(
        messages,
        tokenize=True,
        add_generation_prompt=True,
        enable_thinking=False,
        return_dict=True,
        return_tensors="pt",
    ).to("cuda")
    generated = model.generate(
        **inputs,
        do_sample=False,
        max_new_tokens=max_new_tokens,
        pad_token_id=tokenizer.eos_token_id,
    )
    output_ids = generated[0, inputs.input_ids.shape[1] :]
    return tokenizer.decode(output_ids, skip_special_tokens=True).strip()


def write_workbook(rows: list[dict[str, object]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BF16 与 FP8 对照"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    headers = ["舌象特征", "4B LoRA（BF16）输出", "4B LoRA（FP8）输出"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row["tongue_features"], row["bf16_output"], row["fp8_output"]])

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=3):
        for cell in row:
            cell.alignment = body_alignment
            cell.border = border

    widths = (34, 78, 78)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 28
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 270

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    if not torch.cuda.is_available():
        raise RuntimeError("FP8 inference requires a CUDA GPU.")

    records = get_4b_records(args.source)
    if args.limit is not None:
        if args.limit <= 0:
            raise ValueError("--limit must be positive")
        records = records[: args.limit]
    tokenizer = AutoTokenizer.from_pretrained(args.base_model, local_files_only=True)
    base_model = AutoModelForCausalLM.from_pretrained(
        args.base_model,
        torch_dtype=torch.bfloat16,
        local_files_only=True,
    ).to("cuda")
    # PEFT 0.20 recognises only newer TorchAO releases than the PyTorch 2.7-
    # compatible backend used here.  The base model is still ordinary BF16 at
    # this point, so route LoRA injection through PEFT's normal Linear path;
    # TorchAO is applied only after the adapter has been merged.
    peft_torchao.is_torchao_available = lambda: False
    model = PeftModel.from_pretrained(base_model, args.adapter, local_files_only=True)
    model = model.merge_and_unload()

    bf16_outputs: dict[int, str] = {}
    if args.rerun_bf16:
        print("Running BF16 baseline with the fixed prompts...", flush=True)
        model.eval()
        for record in records:
            messages = prompt_messages(record, args.include_system)
            bf16_outputs[int(record["index"])] = generate(
                model, tokenizer, messages, args.max_new_tokens
            )
            print(f"Completed BF16 {record['index']}/{len(records)}: {record['example_id']}", flush=True)

    print("Quantizing merged model to FP8...", flush=True)
    quantize_(model, float8_dynamic_activation_float8_weight())
    model.eval()

    result_rows: list[dict[str, object]] = []
    for record in records:
        prompt = str(record["dynamic_input"])
        fp8_output = generate(
            model, tokenizer, prompt_messages(record, args.include_system), args.max_new_tokens
        )
        result_rows.append(
            {
                "index": record["index"],
                "example_id": record["example_id"],
                "tongue_features": tongue_features(prompt),
                "bf16_output": bf16_outputs.get(int(record["index"]), str(record["output"])),
                "fp8_output": fp8_output,
            }
        )
        print(f"Completed {record['index']}/{len(records)}: {record['example_id']}", flush=True)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "fp8_results.json").write_text(
        json.dumps(
            {
                "source": str(args.source),
                "base_model": str(args.base_model),
                "adapter": str(args.adapter),
                "quantization": "TorchAO dynamic-activation FP8-weight linear layers",
                "bf16_output_source": (
                    "Generated in this run with the same prompt and generation settings"
                    if args.rerun_bf16
                    else "Saved BF16 output from the source JSON"
                ),
                "generation": {
                    "do_sample": False,
                    "max_new_tokens": args.max_new_tokens,
                    "thinking": False,
                },
                "input_message_roles": ["system", "user"] if args.include_system else ["user"],
                "records": result_rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    write_workbook(result_rows, args.output_dir / "tongue_4b_bf16_fp8_comparison.xlsx")


if __name__ == "__main__":
    main()
