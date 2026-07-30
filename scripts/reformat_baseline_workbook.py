from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tcm_qwen_eval.reporting import write_workbook


def _read_audit_rows(sheet) -> list[dict[str, Any]]:
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, row, strict=True))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in row)
    ]


def _read_legacy_workbook(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=False)
    samples, audit = workbook.worksheets[1], workbook.worksheets[3]
    selected_by_id: dict[str, dict[str, Any]] = {}
    generations: list[dict[str, Any]] = []
    for row in samples.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        model, domain, task, example_id, group_id, _system, user, reference, output = row[:9]
        selected_by_id.setdefault(
            example_id,
            {
                "id": example_id,
                "domain": domain,
                "task": task,
                "group_id": group_id,
                "user": user,
                "reference": reference,
            },
        )
        generations.append(
            {
                "model": model,
                "example_id": example_id,
                "output": output,
                "format_pass": row[9] == "通过",
                "format_note": row[10],
                "output_chars": row[11],
                "latency_seconds": row[12],
                "peak_memory_mib": row[13],
            }
        )
    audit_rows = _read_audit_rows(audit)
    workbook.close()
    return list(selected_by_id.values()), generations, audit_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert the baseline workbook to side-by-side model comparison.")
    parser.add_argument("--input", type=Path, default=Path("artifacts/qwen3_baseline.xlsx"))
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("artifacts/qwen3_baseline_side_by_side.xlsx"),
    )
    args = parser.parse_args()

    selected_rows, generations, audit_rows = _read_legacy_workbook(args.input)
    write_workbook(
        args.output,
        audit_rows,
        selected_rows,
        generations,
        {
            "生成参数": "do_sample=False, max_new_tokens=768, thinking=False",
            "权重精度": "BF16",
            "结果布局": "同一数据组一行；仅保留用户提示词及三个模型的并列输出。",
        },
    )
    print(f"Reformatted {len(selected_rows)} samples / {len(generations)} generations -> {args.output}")


if __name__ == "__main__":
    main()
