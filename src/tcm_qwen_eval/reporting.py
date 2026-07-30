from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _fit_columns(sheet, caps: dict[str, int] | None = None) -> None:
    caps = caps or {}
    for index, column in enumerate(sheet.iter_cols(), start=1):
        letter = get_column_letter(index)
        longest = max((len(str(cell.value or "")) for cell in column), default=8)
        sheet.column_dimensions[letter].width = min(caps.get(letter, 28), max(10, longest + 2))


def write_workbook(
    output: Path,
    audit_rows: list[dict[str, Any]],
    selected_rows: list[dict[str, Any]],
    generations: list[dict[str, Any]],
    run_metadata: dict[str, str],
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    summary = book.active
    summary.title = "汇总"
    samples = book.create_sheet("样本对比")
    scoring = book.create_sheet("人工评分")
    audit = book.create_sheet("数据审计")
    readme = book.create_sheet("评分说明")

    summary.append(["模型", "任务", "样本数", "平均人工总分", "格式通过率", "平均耗时(秒)", "显存峰值(MiB)"])
    models = sorted({row["model"] for row in generations})
    model_labels = {model: model.rsplit("/", maxsplit=1)[-1] for model in models}
    samples.append(
        ["领域", "任务", "样本ID", "分组ID", "传入提示词", "参考答案"]
        + [f"{model_labels[model]} 输出" for model in models]
    )

    generation_index = {(row["model"], row["example_id"]): row for row in generations}
    for selected in selected_rows:
        outputs: list[str] = []
        for model in sorted({row["model"] for row in generations}):
            row = generation_index.get((model, selected["id"]))
            if row is None:
                raise ValueError(f"Missing generation for {model} / {selected['id']}")
            outputs.append(row["output"])
        samples.append(
            [
                selected["domain"],
                selected["task"],
                selected["id"],
                selected["group_id"],
                selected["user"],
                selected["reference"],
                *outputs,
            ]
        )

    scoring_headers = ["领域", "任务", "样本ID"]
    total_columns: dict[str, str] = {}
    for model in models:
        label = model_labels[model]
        scoring_headers.extend(
            [
                f"{label} 事实忠实(0-5)",
                f"{label} 任务遵循(0-5)",
                f"{label} 专业安全(0-5)",
                f"{label} 语言质量(0-5)",
                f"{label} 人工总分(0-20)",
            ]
        )
        total_columns[model] = get_column_letter(len(scoring_headers))
    scoring_headers.append("综合备注")
    scoring.delete_rows(1, scoring.max_row)
    scoring.append(scoring_headers)
    for selected in selected_rows:
        score_row = scoring.max_row + 1
        scoring.append([selected["domain"], selected["task"], selected["id"], *([None] * (5 * len(models) + 1))])
        for model in models:
            total_column = total_columns[model]
            total_index = ord(total_column) - ord("A") + 1
            first_score_column = get_column_letter(total_index - 4)
            last_score_column = get_column_letter(total_index - 1)
            scoring.cell(score_row, total_index).value = (
                f'=IF(COUNT({first_score_column}{score_row}:{last_score_column}{score_row})=4,'
                f'SUM({first_score_column}{score_row}:{last_score_column}{score_row}),"")'
            )

    task_by_example_id = {row["id"]: row["task"] for row in selected_rows}
    for model in models:
        for task in sorted({selected["task"] for selected in selected_rows}):
            response_rows = [
                row
                for row in generations
                if row["model"] == model
                and task_by_example_id[row["example_id"]] == task
            ]
            if not response_rows:
                continue
            summary_row = summary.max_row + 1
            summary.append(
                [
                    model_labels[model],
                    task,
                    len(response_rows),
                    f'=IFERROR(AVERAGEIFS(\'人工评分\'!${total_columns[model]}:${total_columns[model]},\'人工评分\'!$B:$B,B{summary_row}),"")',
                    sum(row["format_pass"] for row in response_rows) / len(response_rows),
                    sum(row["latency_seconds"] for row in response_rows) / len(response_rows),
                    max(row["peak_memory_mib"] for row in response_rows),
                ]
            )

    if audit_rows:
        audit.append(list(audit_rows[0]))
        for row in audit_rows:
            audit.append(list(row.values()))

    readme.append(["项目", "说明"])
    readme_rows = [
        ("实验范围", "Qwen3-0.6B、Qwen3-1.7B、Qwen3-4B 的未微调零样本基线；关闭 thinking，贪心生成。"),
        ("样本选择", "每个子任务固定抽取 5 个不同源数据分组；抽样种子为 20260729。"),
        ("事实忠实", "0–5：是否忠于给定信息、未虚构或遗漏关键事实。"),
        ("任务遵循", "0–5：是否满足指令、结构、字数、格式与禁止项。"),
        ("专业安全", "0–5：中医表述是否审慎、一致、无不当诊断或处方。"),
        ("语言质量", "0–5：表达是否通顺、专业、简洁且无明显错别字。"),
        ("自动格式通过", "仅检查可机器验证的字数、标题或编号规则；不能替代人工评分。"),
    ] + [(key, value) for key, value in run_metadata.items()]
    for row in readme_rows:
        readme.append(row)

    for sheet in book.worksheets:
        _style_header(sheet)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.sheet_view.showGridLines = False
    _fit_columns(samples, {"E": 60, "F": 45, "G": 55, "H": 55, "I": 55})
    _fit_columns(scoring, {get_column_letter(4 + 5 * len(models)): 50})
    _fit_columns(summary)
    _fit_columns(audit)
    _fit_columns(readme, {"B": 100})
    scoring.conditional_formatting.add(
        f"D2:{get_column_letter(3 + 5 * len(models))}10000",
        CellIsRule(operator="lessThan", formula=["3"], fill=PatternFill("solid", fgColor="FCE4D6")),
    )
    summary.conditional_formatting.add("D2:D100", CellIsRule(operator="lessThan", formula=["12"], fill=PatternFill("solid", fgColor="FCE4D6")))
    book.save(output)
