"""Excel reporting for the Qwen3-4B tongue QLoRA evaluation."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from tcm_qwen_eval.dataset import Example
from tcm_qwen_eval.dynamic_input import extract_dynamic_input, task_label

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def _style_body(sheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_tongue_evaluation_workbook(
    output: Path,
    manual_examples: list[Example],
    base_outputs: dict[str, str],
    adapter_outputs: dict[str, str],
    full_predictions: list[dict[str, Any]],
    source_data_dir: Path,
    metadata: dict[str, str],
) -> None:
    """Write a concise manual scoring table plus full held-out format summary."""
    output.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    summary = book.active
    summary.title = "自动汇总"
    results = book.create_sheet("结果表")
    notes = book.create_sheet("说明")

    summary.append(["任务类型", "独立测试样本数", "格式通过数", "格式通过率", "平均输出字数"])
    by_task: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for prediction in full_predictions:
        by_task[prediction["task"]].append(prediction)
    for task, rows in sorted(by_task.items()):
        summary.append(
            [
                task,
                len(rows),
                sum(bool(row["format_pass"]) for row in rows),
                sum(bool(row["format_pass"]) for row in rows) / len(rows),
                sum(int(row["output_chars"]) for row in rows) / len(rows),
            ]
        )
    summary["D2"].number_format = "0.0%"
    summary["D3"].number_format = "0.0%"

    results.append(
        [
            "任务类型",
            "数据源",
            "动态输入数据",
            "Qwen3-4B 基座输出",
            "Qwen3-4B 基座评分",
            "Qwen3-4B QLoRA 输出",
            "Qwen3-4B QLoRA 评分",
            "备注",
        ]
    )
    for example in manual_examples:
        results.append(
            [
                task_label(example),
                str(source_data_dir / f"{example.id.rsplit('/', maxsplit=1)[-1]}.json"),
                extract_dynamic_input(example),
                base_outputs[example.id],
                None,
                adapter_outputs[example.id],
                None,
                None,
            ]
        )

    notes.append(["项目", "说明"])
    notes.append(["人工评分", "评分列留空，供人工按准确性、任务遵循、专业安全和语言质量进行评分。"])
    notes.append(["自动指标", "仅检查长度、句数、编号和 Markdown 等可机械验证规则，不替代人工质量评价。"])
    for key, value in metadata.items():
        notes.append([key, value])

    for sheet in book.worksheets:
        _style_header(sheet)
        _style_body(sheet)
    _set_widths(summary, {"A": 30, "B": 18, "C": 14, "D": 14, "E": 18})
    _set_widths(
        results,
        {"A": 28, "B": 66, "C": 58, "D": 58, "E": 18, "F": 58, "G": 18, "H": 32},
    )
    _set_widths(notes, {"A": 20, "B": 110})
    book.save(output)
