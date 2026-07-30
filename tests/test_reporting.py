from pathlib import Path

from openpyxl import load_workbook

from tcm_qwen_eval.reporting import write_workbook


def test_workbook_contains_scoring_and_summary(tmp_path: Path):
    output = tmp_path / "baseline.xlsx"
    selected = [
        {
            "id": "tongue-analysis/example",
            "domain": "tongue-analysis",
            "task": "tongue_daily_advice",
            "group_id": "tongue-analysis-group",
            "system": "system",
            "user": "user",
            "reference": "reference",
        }
    ]
    generation = [
        {
            "model": "Qwen/Qwen3-0.6B",
            "example_id": "tongue-analysis/example",
            "output": "输出内容",
            "format_pass": True,
            "format_note": "要求 60–80 字且无编号",
            "output_chars": 4,
            "latency_seconds": 1.2,
            "peak_memory_mib": 512.0,
        }
    ]
    write_workbook(
        output,
        [{"领域": "tongue-analysis", "样本数": 1}],
        selected,
        generation,
        {"生成参数": "test"},
    )

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["汇总", "样本对比", "人工评分", "数据审计", "评分说明"]
    comparison = workbook["样本对比"]
    scoring = workbook["人工评分"]
    assert comparison.max_row == 2
    assert comparison.max_column == 7
    assert comparison.cell(2, 5).value == "user"
    assert "system" not in [cell.value for cell in comparison[1]]
    assert scoring.cell(2, 8).value.startswith("=IF(")
    assert workbook["汇总"].max_row == 2
