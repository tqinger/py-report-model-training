from pathlib import Path

import torch
from openpyxl import load_workbook

from tcm_qwen_eval.checkpoints import resolve_resume_checkpoint
from tcm_qwen_eval.dataset import load_examples
from tcm_qwen_eval.tongue_qlora import (
    CausalDataCollator,
    encode_sft_example,
    resolve_model_source,
    split_tongue_examples,
    tongue_format_check,
)
from tcm_qwen_eval.tongue_reporting import write_tongue_evaluation_workbook
from tcm_qwen_eval.training_config import load_tongue_qlora_config


class FakeTokenizer:
    eos_token = "<eos>"
    pad_token_id = 0

    def apply_chat_template(self, messages, tokenize, add_generation_prompt, enable_thinking=False):
        assert not tokenize
        assert add_generation_prompt
        assert enable_thinking is False
        return "PROMPT: " + " ".join(message["content"] for message in messages)

    def __call__(self, text, add_special_tokens=False):
        return {"input_ids": [ord(char) for char in text]}


def test_tongue_qlora_hyperparameters_are_loaded_from_toml_config():
    config = load_tongue_qlora_config(Path("configs/tongue_qlora.toml"))

    assert config.training.learning_rate == 5e-5
    assert config.training.gradient_accumulation_steps == 8
    assert config.training.dataloader_num_workers == 8
    assert config.training.save_steps == 100
    assert config.training.save_total_limit == 1
    assert config.lora.target_modules == (
        "q_proj",
        "k_proj",
        "v_proj",
        "o_proj",
        "gate_proj",
        "up_proj",
        "down_proj",
    )
    assert config.quantization.bnb_4bit_compute_dtype == "bfloat16"


def test_conversations_1_7b_config_uses_a_learning_rate_floor():
    config = load_tongue_qlora_config(Path("configs/tongue_qlora_conversations_1_7b.toml"))

    assert config.training.learning_rate == 3e-5
    assert config.training.lr_scheduler_type == "cosine_with_min_lr"
    assert config.training.lr_scheduler_kwargs == {"min_lr_rate": 0.2}
    assert config.training.per_device_train_batch_size == 8
    assert config.training.gradient_accumulation_steps == 1


def test_new_jsonl_task_configs_match_their_dataset_sequence_lengths():
    expected = {
        "tongue_constitution_50k_qlora.toml": 512,
        "wutai_20_qlora.toml": 2048,
        "holistic_50k_qlora.toml": 1536,
    }

    for filename, max_length in expected.items():
        config = load_tongue_qlora_config(Path("configs") / filename)
        assert config.training.max_length == max_length
        assert config.training.optim == "adamw_torch"
        assert config.training.dataloader_num_workers == 8


def test_tongue_qlora_config_rejects_incompatible_save_and_eval_strategies(tmp_path: Path):
    config_path = tmp_path / "invalid.toml"
    config_path.write_text(
        Path("configs/tongue_qlora.toml")
        .read_text(encoding="utf-8")
        .replace('save_strategy = "steps"', 'save_strategy = "epoch"'),
        encoding="utf-8",
    )

    try:
        load_tongue_qlora_config(config_path)
    except ValueError as error:
        assert "eval_strategy and training.save_strategy" in str(error)
    else:
        raise AssertionError("Expected incompatible save and evaluation strategies to be rejected")


def test_tongue_qlora_config_rejects_non_positive_save_steps(tmp_path: Path):
    config_path = tmp_path / "invalid.toml"
    config_path.write_text(
        Path("configs/tongue_qlora.toml")
        .read_text(encoding="utf-8")
        .replace("save_steps = 100", "save_steps = 0"),
        encoding="utf-8",
    )

    try:
        load_tongue_qlora_config(config_path)
    except ValueError as error:
        assert "training.save_steps must be positive" in str(error)
    else:
        raise AssertionError("Expected non-positive save_steps to be rejected")


def test_resume_from_checkpoint_can_find_and_validate_the_latest_checkpoint(tmp_path: Path):
    older_checkpoint = tmp_path / "checkpoint-100"
    latest_checkpoint = tmp_path / "checkpoint-200"
    older_checkpoint.mkdir()
    latest_checkpoint.mkdir()
    for checkpoint in (older_checkpoint, latest_checkpoint):
        for filename in (
            "trainer_state.json",
            "optimizer.pt",
            "scheduler.pt",
            "adapter_config.json",
            "adapter_model.safetensors",
        ):
            (checkpoint / filename).write_text("{}", encoding="utf-8")

    assert resolve_resume_checkpoint("latest", tmp_path) == str(latest_checkpoint)
    assert resolve_resume_checkpoint(str(older_checkpoint), tmp_path) == str(older_checkpoint)


def test_resume_from_checkpoint_rejects_a_weight_only_directory(tmp_path: Path):
    checkpoint = tmp_path / "checkpoint-100"
    checkpoint.mkdir()

    try:
        resolve_resume_checkpoint(str(checkpoint), tmp_path)
    except SystemExit as error:
        assert "required files" in str(error)
    else:
        raise AssertionError("Expected a weight-only directory to be rejected")


def test_tongue_split_is_disjoint_at_source_group_level():
    splits = split_tongue_examples(load_examples(Path("data")))
    groups = {name: {item.group_id for item in rows} for name, rows in splits.items()}
    assert len(splits["train"]) + len(splits["validation"]) + len(splits["test"]) == 2546
    assert not (groups["train"] & groups["validation"])
    assert not (groups["train"] & groups["test"])
    assert not (groups["validation"] & groups["test"])


def test_model_source_prefers_a_local_huggingface_snapshot(tmp_path: Path):
    snapshot = tmp_path / "hub/models--Qwen--Qwen3-4B/snapshots/revision"
    snapshot.mkdir(parents=True)
    assert resolve_model_source("Qwen/Qwen3-4B", tmp_path) == str(snapshot)


def test_model_source_supports_a_hubless_huggingface_cache(tmp_path: Path):
    snapshot = tmp_path / "models--Qwen--Qwen3-4B/snapshots/revision"
    snapshot.mkdir(parents=True)
    assert resolve_model_source("Qwen/Qwen3-4B", tmp_path) == str(snapshot)


def test_sft_labels_mask_fixed_and_dynamic_prompt_tokens():
    example = next(item for item in load_examples(Path("data")) if item.task == "tongue_daily_advice")
    encoded = encode_sft_example(FakeTokenizer(), example, max_length=4096)
    first_target = next(index for index, label in enumerate(encoded["labels"]) if label != -100)
    assert first_target > 0
    assert encoded["labels"][:first_target] == [-100] * first_target
    assert encoded["labels"][first_target:] == encoded["input_ids"][first_target:]


def test_collator_pads_labels_without_unmasking_prompt():
    batch = CausalDataCollator(0)(
        [
            {"input_ids": [1, 2], "attention_mask": [1, 1], "labels": [-100, 2]},
            {"input_ids": [3], "attention_mask": [1], "labels": [-100]},
        ]
    )
    assert isinstance(batch["input_ids"], torch.Tensor)
    assert batch["labels"].tolist() == [[-100, 2], [-100, -100]]
    assert batch["attention_mask"].tolist() == [[1, 1], [1, 0]]


def test_format_check_has_task_specific_constraints():
    assert tongue_format_check("tongue_daily_advice", "调" * 60)[0]
    assert not tongue_format_check("tongue_daily_advice", "调" * 59)[0]
    assert tongue_format_check("tongue_integrated_analysis", "调" * 50 + "。" + "养" * 49 + "。")[0]
    assert tongue_format_check("tongue_combined_analysis", "分析" * 80 + "建议" * 20)[0]


def test_evaluation_workbook_contains_dynamic_data_and_blank_score_columns(tmp_path: Path):
    examples = [item for item in load_examples(Path("data")) if item.task == "tongue_daily_advice"][:1]
    example = examples[0]
    output = tmp_path / "tongue-evaluation.xlsx"
    write_tongue_evaluation_workbook(
        output,
        examples,
        {example.id: "基座输出"},
        {example.id: "QLoRA 输出"},
        [
            {
                "task": example.task,
                "format_pass": True,
                "output_chars": 64,
            }
        ],
        Path("data/tongue-analysis"),
        {"基座模型": "Qwen/Qwen3-4B"},
    )
    workbook = load_workbook(output)
    results = workbook["结果表"]
    assert [cell.value for cell in results[1]] == [
        "任务类型",
        "数据源",
        "动态输入数据",
        "Qwen3-4B 基座输出",
        "Qwen3-4B 基座评分",
        "Qwen3-4B QLoRA 输出",
        "Qwen3-4B QLoRA 评分",
        "备注",
    ]
    assert "你是一位资深中医" not in results.cell(2, 3).value
    assert results.cell(2, 5).value is None
    assert results.cell(2, 7).value is None
